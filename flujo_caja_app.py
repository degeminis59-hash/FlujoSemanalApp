# -*- coding: utf-8 -*-
"""
Flujo de Caja App - Carga de datos del Requerimiento a las hojas de Proyeccion del Flujo de Caja
Incluye Analisis IA con OpenAI
Ejecutar con: streamlit run flujo_caja_app.py
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import io
import json
import os

# ============================================================
# CONFIGURACION DE MAPEOS
# ============================================================

MAPEOS_FC = [
    {
        "hoja_req": "03_Cuentas por Cobrar",
        "hoja_fc": "Ingresos",
        "fila_inicio": 6,
        "fuente": "Cuentas por Cobrar",
        "campos": [
            ("B", "C"),
            ("D", "J"),
            ("A", "E"),
            ("J", "F"),
            ("E", "G"),
            ("H", "I"),
        ],
        "columnas_formula": ["K"],
    },
    {
        "hoja_req": "05_Ingresos Esperados",
        "hoja_fc": "Ingresos",
        "fila_inicio": None,
        "fuente": "Ingresos esperados",
        "campos": [
            ("A", "E"),
            ("B", "F"),
            ("H", "F"),
            ("C", "J"),
            ("D", "G"),
            ("F", "I"),
        ],
        "columnas_formula": ["K"],
    },
    {
        "hoja_req": "06_Egresos Recurrentes",
        "hoja_fc": "Egresos Recurrentes",
        "fila_inicio": 6,
        "fuente": None,
        "campos": [
            ("A", "D"),
            ("B", "E"),
            ("D", "C"),
            ("J", "F"),
            ("C", "G"),
            ("G", "H"),
            ("H", "I"),
            ("E", "J"),
        ],
        "columnas_formula": ["K"],
    },
    {
        "hoja_req": "07_Planilla",
        "hoja_fc": "Planilla",
        "fila_inicio": 6,
        "fuente": None,
        "campos": [
            ("B", "E"),
            ("D", "D"),
            ("E", "C"),
            ("F", "H"),
            ("H", "G"),
            ("J", "F"),
        ],
        "columnas_formula": ["I"],
    },
]

# ============================================================
# PROMPT DE ANALISIS IA
# ============================================================

PROMPT_ANALISIS = """Eres un analista de tesoreria de una empresa agroexportadora en Peru.
Analiza los datos del Requerimiento de Informacion de Plan de Tesoreria y produce un informe de cobranzas con el siguiente enfoque:

EMPRESA: {empresa}
FECHA DE CORTE: {fecha_corte}
VENTANA DE ANALISIS: 13 semanas desde la fecha de corte

INSTRUCCIONES:
1. Lee TODAS las hojas del archivo y extrae los datos relevantes
2. Identifica las facturas vencidas (fecha de vencimiento/cobro < fecha de corte) y senala cuales NO tienen fecha de cobro confirmada - estas son prioridad de cobranza urgente
3. Para las facturas pendientes, indica en que semana del flujo caen (Semana 1 = primera semana posterior a la fecha de corte)
4. Para los ingresos esperados, indica probabilidad, semana esperada y si son dependientes de confirmacion (comite, SUNAT, etc.)
5. Identifica que semana concentra la mayor entrada de caja por cobranzas
6. Indica si hay riesgo de faltante de caja por desfase entre semanas de cobranza y semanas de pagos criticos

DATOS DEL REQUERIMIENTO:
{datos_requerimiento}

RESPUESTA:
- Formato en espanol
- Usa tablas para mostrar semanas y montos
- Seccion de ALERTAS URGENTES al inicio
- Seccion de RESUMEN SEMANA POR SEMANA (proximas 13 semanas)
- Si falta informacion para calcular algo, indiccalo claramente
"""


# ============================================================
# FUNCIONES DE PROCESAMIENTO
# ============================================================

def combinar_comentarios(valor1, valor2):
    parts = []
    if valor1:
        parts.append(str(valor1))
    if valor2:
        if str(valor2) not in " ".join(parts):
            parts.append(str(valor2))
    return " | ".join(parts) if parts else None


def leer_datos_requerimiento(wb_req, hoja_req):
    if hoja_req not in wb_req.sheetnames:
        return []
    ws = wb_req[hoja_req]
    datos = []
    for row in range(5, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is not None:
                has_data = True
                row_data[cell.column_letter] = val
        if has_data:
            datos.append(row_data)
    return datos


def extraer_datos_para_ia(file):
    """Extrae todos los datos relevantes del requerimiento para enviar al prompt."""
    wb = load_workbook(file, data_only=True)
    resumen = {}

    if "00_Datos Generales" in wb.sheetnames:
        ws = wb["00_Datos Generales"]
        datos_gen = {}
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0] and row[1]:
                datos_gen[str(row[0])] = row[1]
        resumen["Datos_Generales"] = datos_gen

    if "02_Saldos Bancarios" in wb.sheetnames:
        ws = wb["02_Saldos Bancarios"]
        saldos = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                saldos.append({
                    "Banco": row[0], "Cuenta": row[1], "Moneda": row[2],
                    "Saldo": row[3], "Fecha_corte": row[4]
                })
        resumen["Saldos_Bancarios"] = saldos

    if "03_Cuentas por Cobrar" in wb.sheetnames:
        ws = wb["03_Cuentas por Cobrar"]
        cxc = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                cxc.append({
                    "Cliente": row[0], "Documento": row[1],
                    "Fecha_emision": row[2], "Fecha_vencimiento": row[3],
                    "Monto": row[4], "Moneda": row[5],
                    "Estado": row[6], "Probabilidad": row[7],
                    "Prioridad_cobranza": row[8], "Comentario": row[9]
                })
        resumen["Cuentas_por_Cobrar"] = cxc

    if "04_Cuentas por Pagar" in wb.sheetnames:
        ws = wb["04_Cuentas por Pagar"]
        cxp = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                cxp.append({
                    "Proveedor": row[0], "Documento": row[1],
                    "Fecha_emision": row[2], "Fecha_vencimiento": row[3],
                    "Monto": row[4], "Moneda": row[5],
                    "Estado": row[6], "Prioridad_pago": row[7],
                    "Postergable": row[8], "Comentario": row[9]
                })
        resumen["Cuentas_por_Pagar"] = cxp

    if "05_Ingresos Esperados" in wb.sheetnames:
        ws = wb["05_Ingresos Esperados"]
        ing_esp = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                ing_esp.append({
                    "Cliente_fuente": row[0], "Concepto": row[1],
                    "Fecha_esperada": row[2], "Monto": row[3],
                    "Moneda": row[4], "Probabilidad": row[5],
                    "Sustento": row[6], "Comentario": row[7]
                })
        resumen["Ingresos_Esperados"] = ing_esp

    if "06_Egresos Recurrentes" in wb.sheetnames:
        ws = wb["06_Egresos Recurrentes"]
        eg_rec = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                eg_rec.append({
                    "Categoria": row[0], "Concepto": row[1],
                    "Frecuencia": row[2], "Fecha_pago": row[3],
                    "Monto": row[4], "Moneda": row[5],
                    "Prioridad": row[6], "Postergable": row[7],
                    "Comentario": row[9]
                })
        resumen["Egresos_Recurrentes"] = eg_rec

    if "07_Planilla" in wb.sheetnames:
        ws = wb["07_Planilla"]
        planilla = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                planilla.append({
                    "Periodo": row[0], "Area": row[1],
                    "Headcount": row[2], "Concepto": row[3],
                    "Fecha_pago": row[4], "Monto": row[5],
                    "Moneda": row[6], "Prioridad": row[7],
                    "Comentario": row[9]
                })
        resumen["Planilla"] = planilla

    if "08_Deuda y Lineas" in wb.sheetnames:
        ws = wb["08_Deuda y Lineas"]
        deuda = []
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if row[0]:
                deuda.append({
                    "ID": row[0], "Acreedor": row[1],
                    "Producto": row[2], "Moneda": row[3],
                    "Linea_aprobada": row[4], "Saldo_utilizado": row[5],
                    "Disponible": row[6], "Tasa": row[7],
                    "Fecha_vencimiento": row[8], "Estado": row[10]
                })
        resumen["Deuda_y_Lineas"] = deuda

    return resumen


def hacer_analisis_ia(api_key, datos_ia, empresa, fecha_corte):
    try:
        from openai import OpenAI
    except ImportError:
        return "Error: Instala openai con: pip install openai"

    try:
        client = OpenAI(api_key=api_key)
        datos_json = json.dumps(datos_ia, ensure_ascii=False, indent=2, default=str)
        prompt = PROMPT_ANALISIS.format(
            empresa=empresa or "No especificada",
            fecha_corte=fecha_corte or "No especificada",
            datos_requerimiento=datos_json
        )
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Eres un analista financiero especializado en tesoreria de empresas agroexportadoras."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"Error al conectar con OpenAI: {str(e)}"


def aplicar_mapeos_fc(wb_fc, wb_req, mapeos):
    cambios = []
    fila_actual_por_hoja = {}

    for config in mapeos:
        hoja_req = config["hoja_req"]
        hoja_fc = config["hoja_fc"]
        fuente = config.get("fuente")
        campos = config["campos"]
        cols_formula = config.get("columnas_formula", [])

        if hoja_fc not in wb_fc.sheetnames:
            continue

        ws_fc = wb_fc[hoja_fc]

        if config.get("fila_inicio") is not None:
            fila_inicio = config["fila_inicio"]
            fila_actual_por_hoja[hoja_fc] = fila_inicio
        else:
            if hoja_fc not in fila_actual_por_hoja:
                continue
            fila_inicio = fila_actual_por_hoja[hoja_fc]

        datos = leer_datos_requerimiento(wb_req, hoja_req)

        if not datos:
            cambios.append({
                "hoja_req": hoja_req, "hoja_fc": hoja_fc,
                "estado": "sin datos", "registros": 0
            })
            continue

        for idx, row_data in enumerate(datos):
            fila_destino = fila_inicio + idx
            comentario_combinado = None

            for col_req, col_fc in campos:
                if col_fc in cols_formula:
                    continue
                valor = row_data.get(col_req)
                if valor is None:
                    continue
                if col_fc == "F" and hoja_fc == "Ingresos" and hoja_req == "05_Ingresos Esperados":
                    comentario_combinado = combinar_comentarios(comentario_combinado, valor)
                    continue
                ws_fc[f"{col_fc}{fila_destino}"] = valor
                cambios.append({
                    "hoja_req": hoja_req, "hoja_fc": hoja_fc,
                    "celda": f"{col_fc}{fila_destino}",
                    "valor": valor, "registro": idx + 1
                })

            if comentario_combinado:
                ws_fc[f"F{fila_destino}"] = comentario_combinado
                cambios.append({
                    "hoja_req": hoja_req, "hoja_fc": hoja_fc,
                    "celda": f"F{fila_destino}",
                    "valor": comentario_combinado, "registro": idx + 1
                })

            if fuente:
                ws_fc[f"H{fila_destino}"] = fuente
                cambios.append({
                    "hoja_req": hoja_req, "hoja_fc": hoja_fc,
                    "celda": f"H{fila_destino}",
                    "valor": fuente, "registro": idx + 1
                })

            fila_actual_por_hoja[hoja_fc] = fila_destino + 1

        cambios.append({
            "hoja_req": hoja_req, "hoja_fc": hoja_fc,
            "estado": "ok", "registros": len(datos)
        })

    return cambios


def generar_resumen(cambios):
    resumen = {}
    for c in cambios:
        if c.get("estado") == "ok":
            key = f"{c['hoja_req']} -> {c['hoja_fc']}"
            resumen[key] = c["registros"]
    return resumen


# ============================================================
# INTERFAZ STREAMLIT
# ============================================================

def main():
    st.set_page_config(
        page_title="Flujo de Caja - Carga Proyeccion",
        page_icon="📈",
        layout="wide"
    )

    st.title("📈 Carga de Proyecciones al Flujo de Caja")
    st.markdown("Transferencia de datos del **Requerimiento** a las hojas de **proyeccion** del Flujo de Caja Semanal.")
    st.markdown("---")

    # Sidebar
    st.sidebar.header("Configuracion")
    st.sidebar.markdown("""
    | Requerimiento | -> | Flujo de Caja |
    |---|---|---|
    | 03_Cuentas por Cobrar | -> | Ingresos |
    | 05_Ingresos Esperados | -> | Ingresos |
    | 06_Egresos Recurrentes | -> | Egresos Recurrentes |
    | 07_Planilla | -> | Planilla |
    """)
    st.sidebar.markdown("""
    **Columnas automaticas (no se escriben):**
    - `K` en Ingresos y Egresos Recurrentes (Semana flujo)
    - `I` en Planilla (Semana flujo)
    """)

    # Session state
    for key in ['fc_procesado', 'fc_nombre', 'cambios_guardados', 'resultado_ia']:
        if key not in st.session_state:
            st.session_state[key] = None

    # Upload archivos
    col1, col2 = st.columns(2)

    with col1:
        st.header("1. Archivo Requerimiento")
        requerimiento_file = st.file_uploader(
            "Subir archivo Requerimiento de la empresa",
            type=["xlsx", "xls"],
            key="req_uploader"
        )

    with col2:
        st.header("2. Plantilla Flujo de Caja")
        flujo_caja_file = st.file_uploader(
            "Subir plantilla Flujo de Caja Semanal",
            type=["xlsx", "xls"],
            key="fc_uploader"
        )

    st.markdown("---")

    # Seccion de analisis IA
    st.header("3. Analisis IA")
    st.info("Usa OpenAI para analizar automaticamente el requerimiento de tesoreria.")

    api_key = os.environ.get("OPENAI_API_KEY", "")

    col_api, col_btn = st.columns([2, 1])
    with col_api:
        if api_key:
            st.success("API Key de OpenAI configurada correctamente")
        else:
            st.warning("No se encontro OPENAI_API_KEY. Configurala como variable de entorno.")

    with col_btn:
        st.markdown("")
        analizar_disabled = not (requerimiento_file and api_key)
        if st.button("🤖 Analizar con IA", type="primary", disabled=analizar_disabled, key="analisis_ia_btn"):
            if not requerimiento_file:
                st.error("Sube el Requerimiento primero")
            elif not api_key:
                st.error("Configura la variable OPENAI_API_KEY primero")
            else:
                with st.spinner("Analizando con IA..."):
                    try:
                        datos_ia = extraer_datos_para_ia(requerimiento_file)
                        empresa = datos_ia.get("Datos_Generales", {}).get("Empresa", "")
                        fecha_corte = datos_ia.get("Datos_Generales", {}).get("Fecha de corte de informacion", "")
                        resultado = hacer_analisis_ia(api_key, datos_ia, empresa, fecha_corte)
                        st.session_state['resultado_ia'] = resultado
                    except Exception as e:
                        st.error(f"Error: {str(e)}")

    # Mostrar resultado del analisis
    if st.session_state.get('resultado_ia'):
        st.markdown("---")
        st.markdown("### Resultado del Analisis")
        st.markdown(st.session_state['resultado_ia'])

    st.markdown("---")

    # Previsualizar
    if requerimiento_file and flujo_caja_file:
        if st.button("Previsualizar datos", key="preview_btn"):
            with st.spinner("Analizando archivos..."):
                try:
                    wb_req = load_workbook(requerimiento_file, data_only=True)
                    wb_fc = load_workbook(flujo_caja_file, data_only=True)
                    st.success("Archivos cargados correctamente")
                    for config in MAPEOS_FC:
                        hoja_req = config["hoja_req"]
                        hoja_fc = config["hoja_fc"]
                        datos = leer_datos_requerimiento(wb_req, hoja_req)
                        if not datos:
                            continue
                        with st.expander(f"📋 {hoja_req} -> **{hoja_fc}** ({len(datos)} registros)"):
                            st.dataframe(pd.DataFrame(datos), use_container_width=True)
                except Exception as e:
                    st.error(f"Error al previsualizar: {e}")

    st.markdown("---")

    # Aplicar cambios
    st.header("4. Aplicar cambios")

    aplicar_disabled = not (requerimiento_file and flujo_caja_file)

    if st.button("Transferir datos al Flujo de Caja", type="primary", disabled=aplicar_disabled, key="transfer_btn"):
        if not requerimiento_file:
            st.error("Sube el archivo de Requerimiento primero")
        elif not flujo_caja_file:
            st.error("Sube la plantilla de Flujo de Caja primero")
        else:
            with st.spinner("Procesando..."):
                try:
                    wb_fc = load_workbook(flujo_caja_file)
                    wb_req = load_workbook(requerimiento_file, data_only=True)
                    cambios = aplicar_mapeos_fc(wb_fc, wb_req, MAPEOS_FC)
                    output = io.BytesIO()
                    wb_fc.save(output)
                    output.seek(0)
                    ts = datetime.now().strftime("%Y%m%d_%H%M")
                    st.session_state['fc_procesado'] = output.getvalue()
                    st.session_state['fc_nombre'] = f"Flujo_Caja_Proyeccion_{ts}.xlsx"
                    st.session_state['cambios_guardados'] = cambios
                    resumen = generar_resumen(cambios)
                    st.success(f"Transferencia completada - {len(cambios)} celdas escritas")
                    if resumen:
                        st.markdown("**Registros transferidos:**")
                        for k, v in resumen.items():
                            st.write(f"  - {k}: **{v}** registros")
                except Exception as e:
                    st.error(f"Error: {str(e)}")
                    import traceback
                    st.text(traceback.format_exc())

    # Descargar
    if st.session_state['fc_procesado']:
        st.markdown("---")
        st.download_button(
            label="Descargar Flujo de Caja con proyecciones",
            data=st.session_state['fc_procesado'],
            file_name=st.session_state['fc_nombre'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        cambios = st.session_state.get('cambios_guardados', [])
        datos_changes = [c for c in cambios if "celda" in c]
        if datos_changes:
            with st.expander("Ver celdas escritas"):
                for c in datos_changes[:20]:
                    st.write(f"  `{c['hoja_fc']}!{c['celda']}` = {c['valor']}")
                if len(datos_changes) > 20:
                    st.write(f"  _... y {len(datos_changes) - 20} celdas mas_")

    st.markdown("---")
    st.info("""
    **Nota:** Las columnas de formula (Semana flujo) se calculan automaticamente en Excel.
    Si necesitas que se actualicen, abre el archivo en Excel y guarda - o presiona F9.
    """)


if __name__ == "__main__":
    main()
